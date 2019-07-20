import os
import openpyxl as excel
from openpyxl.worksheet import worksheet
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from . import WaimokuUser


class WaimokuClient:
    __target_sheet_name = "入力フォーマット"

    def save_to_file(self, user_list: [WaimokuUser], save_filename="event_participantsList.xlsx"):
        """ユーザ情報一覧をLODGE提出用のフォーマットのXLSファイルに保存する

        Arguments:
            user_list {[WaimokuUser]} -- ユーザ情報一覧

        Keyword Arguments:
            save_filename {str} -- LODGE提出用のフォーマットのXLSファイルの名前 (default: {"event_participantsList.xlsx"})
        """
        # ブックの新規作成
        wb = excel.load_workbook(os.path.dirname(os.path.abspath(__file__)) + "/res/event_visitorList.xlsx")
        ws = wb[self.__target_sheet_name]
        self.__del_rows(ws=ws, idx=3, amount=3)
        for index, user in enumerate(user_list):
            self.__write(ws=ws, key="A{0}".format(index + 3), value=index+1, font_size=9)
            self.__write(ws=ws, key="C{0}".format(index + 3), value=user.full_name, font_size=11)
            self.__write(ws=ws, key="D{0}".format(index + 3), value=user.assign, font_size=11)
        wb.save(save_filename)

    def __del_rows(self, ws: worksheet, idx: int, amount=1):
        """指定した範囲の行を削除する

        Arguments:
            ws {worksheet} -- 対象のワークシート
            idx {int} -- 対象インデックス

        Keyword Arguments:
            amount {int} -- 削除する範囲。指定した行数分削除される (default: {1})
        """
        ws.delete_rows(idx, amount)

    def __write(self, ws, key: str, value, fill: PatternFill = PatternFill(), border=Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000')), font_size: int = 9):
        """ワークシートの指定したキーにデータを書き込む

        Arguments:
            ws {} -- 対象のワークシート
            key {str} -- キー
            value {} -- 書き込む内容

        Keyword Arguments:
            fill {PatternFill} -- 色の指定 (default: {PatternFill()})
            border {Border} -- 罫線情報(default: {Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))})
            font_size {int} -- フォントサイズ (default: {9})
        """
        ws[key] = value
        ws[key].fill = fill
        ws[key].border = border
        ws[key].font = Font(size=font_size)
